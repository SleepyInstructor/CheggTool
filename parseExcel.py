from openpyxl import load_workbook


def parseChegg(filename: str):
    columns = ["Question ID", "Question Date", "Answer Date", "Asker User ID", "Asker First Name", "Asker Last Name","Asker Email ID",
               "Asker IP Address", "Asker School Name", "Question", "Answer", ]

    ws_name = "Asker Details"
    ws_name_alt = "Asker details"
    wb = load_workbook(filename)

    if ws_name in wb:
        chegg_ws = wb[ws_name]
    elif ws_name_alt in wb:
        chegg_ws = wb[ws_name_alt]
    else:
        return []

    i = 2
    table = []
    while not chegg_ws.cell(row=i, column=1).value is None:
        data = {}
        for col in range(1, 12):
            data[columns[col - 1]] = chegg_ws.cell(row=i, column=col).value

        table.append(data)
        i += 1

    return table
